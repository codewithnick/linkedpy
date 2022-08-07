import pickle
import openpyxl
from selenium import webdriver
import time
import re
import traceback
from datetime import datetime
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import json
import random
from proxymaker import *
def createexcelifnotexists():
    filepath="database.xlsx"
    if not os.path.isfile(filepath):
        #if not exists file
        wb = openpyxl.Workbook()
        wb.save(filepath)
def readfromexcel(sheet_name,row,col):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    return sheet_obj.cell(row = row, column = col).value

def writetoexcel(sheet_name,row,col,value):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    sheet_obj.cell(row = row, column = col).value= value
    wb_obj.save("database.xlsx")
def getmaxrow(sheet_name):
    createexcelifnotexists()
    wb_obj = openpyxl.load_workbook("database.xlsx")
    try:    
        sheet_obj = wb_obj[sheet_name]
    except KeyError:
        wb_obj.create_sheet(sheet_name)
        sheet_obj = wb_obj[sheet_name]
    return sheet_obj.max_row

class Browser:
    def __init__(self,username,password):
        chrome_options = Options()
        prefs = {
            "download_restrictions": 3,
            "download.open_pdf_in_system_reader": False,
            "download.prompt_for_download": True,
            "download.default_directory": "/dev/null",
            "plugins.always_open_pdf_externally": False,
            "profile.managed_default_content_settings.images": 2
        }
        chrome_options.add_experimental_option(
            "prefs", prefs
        )
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument('log-level=3')
        prefs = {
            "download_restrictions": 3,
        }
        chrome_options.add_experimental_option(
            "prefs", prefs
        )
        PROXY = get_random_proxy()
        #chrome_options.add_argument('--proxy-server=%s' % PROXY)
        driver = webdriver.Chrome('chromedriver',options=chrome_options)
        self.driver=driver          
        self.setmobileview()
        self.username=username
        self.password=password
        self.login()
    def setmobileview(self):
         # Apple iPhone X:      375, 812
        # Apple iPhone XS Max: 414, 896
        try:
            self.driver.set_window_size(414, 896)
        except :
            print("Unexpected alert on resizing web driver!\n\t")
            
    def login(self):
        self.driver.get("https://www.linkedin.com")
        try:
            print("trying to load cookie if available")
            self.loadcookie()
            self.driver.get("https://www.linkedin.com")
            return
        except:
            print("some problem with cookie or its not available")
            traceback.print_exc()
        time.sleep(10)
        self.driver.get("https://www.linkedin.com/login")
        self.driver.find_element_by_id('username').send_keys(self.username)
        self.driver.find_element_by_id('password').send_keys(self.password)
        self.driver.find_element_by_id('password').send_keys("\n")
        try:
            self.driver.find_elements_by_xpath("//*[contains(text(),'Skip']").click()
        except:
            pass
        self.savecookies()
    def loadcookie(self):
        print("loading cookie")
        cookies = pickle.load(open("cookies.pkl", "rb"))
        print(cookies)
        self.driver.add_cookie(cookies)
        print('loaded cookie')
    def savecookies(self):
        print("saving cookie")
        time.sleep(10)
        cookies=self.driver.get_cookies()
        for cookie in cookies:
            if(cookie['name']=='li_at'):
                cookie['domain']='.linkedin.com'
                x={
                'name': 'li_at',
                'value': cookie['value'],
                'domain': '.linkedin.com'
                }
                break
        pickle.dump(x , open("cookies.pkl","wb"))

        print('cookies saved')
    def leave(self):
        print(" killing the browser ")
        self.driver.quit()